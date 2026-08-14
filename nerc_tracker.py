import hashlib
import json
import os
import re
import smtplib
from collections import defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from email.message import EmailMessage
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

EXCEL_URL = "https://www.nerc.com/globalassets/align-reports/one-stop-shop.xlsx"
PROFILE_PATH = Path("docs/data/profile.json")
STATE_PATH = Path("state/nerc-records.json")
DATA_PATH = Path("docs/data/nerc.json")
HISTORY_PATH = Path("docs/data/nerc-history.json")
MAX_WEB_CHANGES = 500
MAX_EMAIL_CHANGES = 25

GMAIL_USER = os.environ.get("GMAIL_USER")
GMAIL_PASS = os.environ.get("GMAIL_PASS")
RECIPIENTS = [x.strip() for x in os.environ.get(
    "RECIPIENTS", "mikep@mcphersonpower.com,tommys@mcphersonpower.com"
).split(",") if x.strip()]

STANDARD_RE = re.compile(r"\b([A-Z]{3})-(\d{3})(?:-[A-Za-z0-9.]+)?\b")
HEADER_WORDS = (
    "standard", "requirement", "version", "status", "enforcement", "effective",
    "implementation", "subject", "title", "project", "date", "retirement",
)
IDENTITY_WORDS = (
    "standard", "requirement", "requirement number", "section", "subject",
    "title", "project", "version", "criterion", "reference",
)
VOLATILE_WORDS = (
    "status", "enforcement", "effective", "implementation", "date", "retirement",
    "inactive", "mandatory", "future enforcement", "notes", "comment",
)

def load_json(path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return default

def load_profile():
    profile = load_json(PROFILE_PATH, {})
    tracked = set(profile.get("tracked_standards", []))
    if not tracked:
        raise RuntimeError("No tracked standards are configured in docs/data/profile.json")
    return profile, tracked

def download_excel():
    response = requests.get(
        EXCEL_URL, timeout=90,
        headers={"User-Agent": "BPU-NERC-Regulatory-Control-Room/2.0"},
    )
    response.raise_for_status()
    return response.content

def value_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()

def normalize(text):
    return re.sub(r"[^a-z0-9]+", " ", value_text(text).lower()).strip()

def extract_standards(values):
    found = set()
    for value in values:
        for match in STANDARD_RE.finditer(value_text(value).upper()):
            found.add(f"{match.group(1)}-{match.group(2)}")
    return sorted(found)

def unique_headers(values):
    result, seen = [], defaultdict(int)
    for idx, value in enumerate(values, start=1):
        name = value_text(value) or f"Column {idx}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name} ({seen[name]})"
        result.append(name)
    return result

def find_header_row(rows):
    best_index, best_score = 0, -1
    for idx, row in enumerate(rows[:30]):
        texts = [normalize(v) for v in row if value_text(v)]
        if len(texts) < 3:
            continue
        keyword_hits = sum(1 for text in texts for word in HEADER_WORDS if word in text)
        score = keyword_hits * 5 + min(len(texts), 20)
        if score > best_score:
            best_score, best_index = score, idx
    return best_index

def stable_identity(fields, standards):
    selected = []
    for header, value in fields.items():
        if not value:
            continue
        h = normalize(header)
        if any(word in h for word in VOLATILE_WORDS):
            continue
        if any(word in h for word in IDENTITY_WORDS):
            selected.append(f"{h}={normalize(value)}")
    if not selected:
        for header, value in list(fields.items())[:6]:
            h = normalize(header)
            if value and not any(word in h for word in VOLATILE_WORDS):
                selected.append(f"{h}={normalize(value)}")
    standard_part = ",".join(standards) if standards else "unclassified"
    identity_text = standard_part + "|" + "|".join(selected[:8])
    key = hashlib.sha1(identity_text.encode("utf-8")).hexdigest()[:20]
    return key, identity_text

def parse_workbook(path, tracked):
    workbook = load_workbook(path, read_only=True, data_only=False)
    records, sheet_meta = [], []
    total_populated = 0

    for sheet in workbook.worksheets:
        raw_rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        if not raw_rows:
            continue
        header_idx = find_header_row(raw_rows)
        headers = unique_headers(raw_rows[header_idx])
        previous_standards = []

        sheet_record_start = len(records)
        for excel_row, values in enumerate(raw_rows[header_idx + 1:], start=header_idx + 2):
            text_values = [value_text(v) for v in values]
            populated = sum(1 for v in text_values if v)
            total_populated += populated
            if populated == 0:
                previous_standards = []
                continue

            fields = {
                headers[i] if i < len(headers) else f"Column {i + 1}": value
                for i, value in enumerate(text_values) if value
            }
            standards = extract_standards(text_values)
            if standards:
                previous_standards = standards
            elif previous_standards:
                standards = previous_standards

            key, identity_text = stable_identity(fields, standards)
            primary = standards[0] if standards else ""
            records.append({
                "key": key,
                "identity_text": identity_text,
                "sheet": sheet.title,
                "source_row": excel_row,
                "standards": standards,
                "standard": primary,
                "family": primary.split("-")[0] if primary else "OTHER",
                "tracked": any(s in tracked for s in standards),
                "fields": fields,
            })

        sheet_meta.append({
            "name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "records": len(records) - sheet_record_start,
        })

    workbook.close()
    return records, sheet_meta, total_populated

def similarity(old, new):
    if old.get("standard") and new.get("standard") and old["standard"] != new["standard"]:
        return 0.0
    return SequenceMatcher(
        None, old.get("identity_text", ""), new.get("identity_text", "")
    ).ratio()

def severity_for(change):
    if not change.get("tracked"):
        return "info"
    text = " ".join([
        change.get("field", ""), change.get("old", ""), change.get("new", "")
    ]).lower()
    high_terms = (
        "enforcement", "mandatory", "inactive", "effective date",
        "implementation", "retirement", "subject to future enforcement",
    )
    moderate_terms = (
        "requirement", "applicability", "version", "status", "violation",
        "approval", "ballot",
    )
    if any(term in text for term in high_terms):
        return "high"
    if any(term in text for term in moderate_terms):
        return "moderate"
    return "informational"

def field_changes(old, new):
    changes = []
    for header in sorted(set(old.get("fields", {})) | set(new.get("fields", {}))):
        old_value = value_text(old.get("fields", {}).get(header, ""))
        new_value = value_text(new.get("fields", {}).get(header, ""))
        if old_value == new_value:
            continue
        change_type = "modified"
        if not old_value and new_value:
            change_type = "added"
        elif old_value and not new_value:
            change_type = "removed"
        standard = new.get("standard") or old.get("standard") or "Unclassified"
        item = {
            "standard": standard,
            "family": standard.split("-")[0] if "-" in standard else "OTHER",
            "sheet": new.get("sheet") or old.get("sheet", ""),
            "field": header,
            "old": old_value,
            "new": new_value,
            "type": change_type,
            "tracked": bool(new.get("tracked") or old.get("tracked")),
            "source_row": new.get("source_row") or old.get("source_row"),
        }
        item["severity"] = severity_for(item)
        changes.append(item)
    return changes

def compare_records(previous, current):
    old_by_key, new_by_key = defaultdict(list), defaultdict(list)
    for record in previous:
        old_by_key[record["key"]].append(record)
    for record in current:
        new_by_key[record["key"]].append(record)

    matched, old_unmatched, new_unmatched = [], [], []
    for key in set(old_by_key) | set(new_by_key):
        old_group, new_group = old_by_key.get(key, []), new_by_key.get(key, [])
        count = min(len(old_group), len(new_group))
        matched.extend(zip(old_group[:count], new_group[:count]))
        old_unmatched.extend(old_group[count:])
        new_unmatched.extend(new_group[count:])

    candidates = []
    for oi, old in enumerate(old_unmatched):
        for ni, new in enumerate(new_unmatched):
            if old.get("sheet") != new.get("sheet"):
                continue
            score = similarity(old, new)
            if score >= 0.72:
                candidates.append((score, oi, ni))
    candidates.sort(reverse=True)

    used_old, used_new = set(), set()
    for score, oi, ni in candidates:
        if oi in used_old or ni in used_new:
            continue
        used_old.add(oi)
        used_new.add(ni)
        matched.append((old_unmatched[oi], new_unmatched[ni]))

    changes = []
    for old, new in matched:
        changes.extend(field_changes(old, new))

    for idx, record in enumerate(old_unmatched):
        if idx in used_old:
            continue
        standard = record.get("standard") or "Unclassified"
        item = {
            "standard": standard,
            "family": record.get("family", "OTHER"),
            "sheet": record.get("sheet", ""),
            "field": "Record",
            "old": "Record present",
            "new": "Record removed",
            "type": "record_removed",
            "tracked": record.get("tracked", False),
            "source_row": record.get("source_row"),
        }
        item["severity"] = severity_for(item)
        changes.append(item)

    for idx, record in enumerate(new_unmatched):
        if idx in used_new:
            continue
        standard = record.get("standard") or "Unclassified"
        item = {
            "standard": standard,
            "family": record.get("family", "OTHER"),
            "sheet": record.get("sheet", ""),
            "field": "Record",
            "old": "Record absent",
            "new": "Record added",
            "type": "record_added",
            "tracked": record.get("tracked", False),
            "source_row": record.get("source_row"),
        }
        item["severity"] = severity_for(item)
        changes.append(item)
    return changes

def summarize_standards(records, changes, profile):
    record_counts = defaultdict(int)
    for record in records:
        for standard in record.get("standards", []):
            record_counts[standard] += 1

    change_counts = defaultdict(lambda: {
        "total": 0, "high": 0, "moderate": 0, "informational": 0
    })
    for change in changes:
        if not change.get("tracked"):
            continue
        standard = change.get("standard", "")
        bucket = change_counts[standard]
        bucket["total"] += 1
        sev = change.get("severity", "informational")
        if sev in bucket:
            bucket[sev] += 1

    standards = []
    for standard in profile.get("tracked_standards", []):
        counts = change_counts[standard]
        standards.append({
            "standard": standard,
            "family": standard.split("-")[0],
            "record_count": record_counts[standard],
            "changes": counts["total"],
            "high": counts["high"],
            "moderate": counts["moderate"],
            "informational": counts["informational"],
            "status": "attention" if counts["high"] else (
                "changed" if counts["total"] else "normal"
            ),
        })

    families = []
    for family in profile.get("families", []):
        entries = [s for s in standards if s["family"] == family["code"]]
        families.append({
            "code": family["code"],
            "name": family["name"],
            "standards": len(entries),
            "changes": sum(s["changes"] for s in entries),
            "high": sum(s["high"] for s in entries),
        })
    return standards, families

def write_dashboard(payload, history, records):
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    HISTORY_PATH.write_text(json.dumps(history[:180], indent=2, ensure_ascii=False), encoding="utf-8")
    STATE_PATH.write_text(json.dumps({"records": records}, indent=2, ensure_ascii=False), encoding="utf-8")

def send_email(subject, body):
    if not GMAIL_USER or not GMAIL_PASS:
        raise RuntimeError("GMAIL_USER or GMAIL_PASS is not set.")
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = GMAIL_USER
    message["To"] = ", ".join(RECIPIENTS)
    message.set_content(body)
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=60) as server:
        server.starttls()
        server.login(GMAIL_USER, GMAIL_PASS)
        server.send_message(message)

def main():
    profile, tracked = load_profile()
    now_utc = datetime.now(timezone.utc)
    now_ct = now_utc.astimezone(ZoneInfo("America/Chicago"))
    today = now_ct.strftime("%Y-%m-%d")

    temp_path = Path("current_snapshot.xlsx")
    temp_path.write_bytes(download_excel())
    records, sheets, populated_cells = parse_workbook(temp_path, tracked)
    temp_path.unlink(missing_ok=True)

    previous_records = load_json(STATE_PATH, {}).get("records", [])
    baseline = not bool(previous_records)
    changes = [] if baseline else compare_records(previous_records, records)

    tracked_changes = [c for c in changes if c.get("tracked")]
    other_changes = [c for c in changes if not c.get("tracked")]
    high = sum(1 for c in tracked_changes if c["severity"] == "high")
    moderate = sum(1 for c in tracked_changes if c["severity"] == "moderate")

    if baseline:
        status_text = "Semantic baseline established"
        brief = (
            f"Version 2 established a record-aware NERC baseline across {len(records):,} records. "
            "Future checks compare records by identity instead of Excel row or cell position."
        )
    elif tracked_changes:
        status_text = "Tracked changes detected"
        brief = (
            f"{len(tracked_changes):,} changes affect BPU-tracked standards "
            f"({high} high priority, {moderate} moderate). "
            f"{len(other_changes):,} additional changes were outside the tracked-standard list."
        )
    else:
        status_text = "No tracked changes"
        brief = (
            f"No changes were detected in BPU's tracked standards. "
            f"{len(other_changes):,} changes were detected elsewhere in the NERC workbook."
            if other_changes else
            "No semantic record changes were detected in the NERC workbook."
        )

    standard_summary, family_summary = summarize_standards(records, changes, profile)
    history = load_json(HISTORY_PATH, [])
    history = [{
        "date": now_ct.strftime("%b %d, %Y %I:%M %p CT"),
        "tracked_changes": len(tracked_changes),
        "other_changes": len(other_changes),
        "high": high,
        "status_text": status_text,
    }] + history

    payload = {
        "schema_version": 2,
        "status": "ok",
        "message": status_text,
        "checked_at": now_utc.isoformat(),
        "checked_at_display": now_ct.strftime("%b %d, %Y %I:%M %p"),
        "source_url": EXCEL_URL,
        "baseline": baseline,
        "summary": {
            "tracked_standards": len(tracked),
            "registered_functions": profile.get("registered_functions", []),
            "sheets_checked": len(sheets),
            "records_reviewed": len(records),
            "cells_reviewed": populated_cells,
            "tracked_changes": len(tracked_changes),
            "other_changes": len(other_changes),
            "high": high,
            "moderate": moderate,
            "informational": max(0, len(tracked_changes) - high - moderate),
        },
        "brief": brief,
        "changes": tracked_changes[:MAX_WEB_CHANGES],
        "other_changes": other_changes[:100],
        "changes_truncated": len(tracked_changes) > MAX_WEB_CHANGES,
        "standards": standard_summary,
        "families": family_summary,
        "sheets": sheets,
        "history": history[:45],
        "health": {
            "nerc_feed": "ok",
            "semantic_compare": "ok",
            "email": "pending",
        },
    }
    write_dashboard(payload, history, records)

    details = []
    for item in tracked_changes[:MAX_EMAIL_CHANGES]:
        details.append(
            f"{item['standard']} | {item['field']} | {item['severity'].upper()}\n"
            f"  Previous: {item['old'] or '(blank)'}\n"
            f"  Current:  {item['new'] or '(blank)'}"
        )
    if len(tracked_changes) > MAX_EMAIL_CHANGES:
        details.append(
            f"...and {len(tracked_changes) - MAX_EMAIL_CHANGES} additional tracked changes. "
            "See the Control Room for detail."
        )

    subject = f"[BPU NERC Control Room] {status_text} - {today}"
    body = brief + (("\n\n" + "\n\n".join(details)) if details else "")
    body += "\n\nControl Room: https://mcpbpunerc.github.io/nerc-one-stop-shop-tracker/"

    try:
        send_email(subject, body)
        payload["health"]["email"] = "ok"
    except Exception as exc:
        payload["health"]["email"] = "error"
        payload["health"]["email_error"] = type(exc).__name__
        print(f"WARNING: email delivery failed: {exc}")

    DATA_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

if __name__ == "__main__":
    main()
